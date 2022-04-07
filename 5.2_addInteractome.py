#!/usr/bin/python

# manojmw
# 15 Mar, 2022

import argparse, sys
import openpyxl as xl
import scipy.stats as stats
import re
import gzip
import logging

###########################################################

# Parses tab-seperated canonical transcripts file
# Required columns are: 'ENSG' and 'GENE' (can be in any order,
# but they MUST exist)
#
# Returns a dictionary:
# - Key -> ENSG
# - Value -> Gene
def ENSG_Gene(inCanonicalFile):

    logging.info("starting to run")

    # Dictionary to store ENSG & Gene data
    ENSG_Gene_dict = {}

    # Opening canonical transcript file (gzip or non-gzip)
    try:
        if inCanonicalFile.endswith('.gz'):
            Canonical_File = gzip.open(inCanonicalFile, 'rt')
        else:
            Canonical_File = open(inCanonicalFile)
    except IOError:
        logging.error("At Step 5.2_addInteractome - Failed to read the Canonical transcript file: %s" % inCanonicalFile)
        sys.exit()

    Canonical_header_line = Canonical_File.readline() # Grabbing the header line

    Canonical_header_fields = Canonical_header_line.split('\t')

    # Check the column headers and grab indexes of our columns of interest
    (ENSG_index, Gene_index) = (-1,-1)

    for i in range(len(Canonical_header_fields)):
        if Canonical_header_fields[i] == 'ENSG':
            ENSG_index = i
        elif Canonical_header_fields[i] == 'GENE':
            Gene_index = i

    if not ENSG_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'ENSG' in the file: %s \n" % inCanonicalFile)
        sys.exit()
    elif not Gene_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'GENE' in the file: %s \n" % inCanonicalFile)
        sys.exit()
    # else grabbed the required column indexes -> PROCEED

    # Data lines
    for line in Canonical_File:
        line = line.rstrip('\n')
        CanonicalTranscripts_fields = line.split('\t')

        # Key -> ENSG
        # Value -> Gene

        ENSG_Gene_dict[CanonicalTranscripts_fields[ENSG_index]] = CanonicalTranscripts_fields[Gene_index]

    # Closing the file
    Canonical_File.close()

    return ENSG_Gene_dict

###########################################################

# Parses the candidateGenes file in .xlsx format
# Required columns are: 'Gene' & 'pathologyID' 
# (can be in any order, but they MUST exist)
# Also parses {ENSG_Gene_dict} returned
# by the function: ENSG_Gene
# Maps Candidate Gene names to ENSG
#
# Returns a dictionary
# - Key: ENSG of Candidate Gene
# - Value: list of pathology(s)
def CandidateGeneParser(inCandidateFile, ENSG_Gene_dict):
    
    # Input - List of candidate gene file(s)
    candidate_files = inCandidateFile

    # Dictionary to store candidate genes
    # and associated data
    # Key: Candidate Gene
    # Value: list of pathology(s)
    CandidateGene_dict = {}

    # Data lines
    for file in candidate_files:

        # Creating a workbook object 
        wb_obj = xl.load_workbook(file)

        # Creating a sheet object from the active attribute
        sheet_obj = wb_obj.active

        # Dictionary to store Candidate Gene and pathology data
        # Key -> rowindex of pathologyID
        # Value -> pathologyID
        Gene_patho_dict = {}
        
        # Iterating over col cells and checking if any header 
        # in the the header line matches our header of interest
        for header_cols in sheet_obj.iter_cols(1, sheet_obj.max_column):
            if header_cols[0].value == "pathologyID":
                for patho_field in header_cols[1:]:
                    # Skip empty fields
                    if patho_field.value == None:
                        pass
                    else:
                        Gene_patho_dict[patho_field.row] = patho_field.value

        # Grabbing gene names
        # Replacing the key (rowindex of pathologyID) in Gene_patho_dict
        # with a our new key (Gene_identifier)
        # Gene_identifier -> Gene name & rowindex of Gene name seperated by an '_'
        # This is to make sure that we do not replace the existsing gene and patho 
        # if the same gene is associated with a different pathology (in a given file,
        # row index will be unique)
        # Using a new for-loop because the keys in Gene_patho_dict will 
        # not be defined until we exit for loop
        # If key is not defined, then we cannot replace the old key with
        # our new Gene_identifier using the same row index
        for header_cols in sheet_obj.iter_cols(1, sheet_obj.max_column):
            if header_cols[0].value == "Gene":
                for Gene_field in header_cols[1:]:
                    # Skip empty fields
                    if Gene_field.value == None:
                        pass
                    else:
                        # Replacing the key in Gene_patho_dict with our new key (Gene_identifier)
                        Gene_patho_dict[Gene_field.value + '_' + str(Gene_field.row)] = Gene_patho_dict.pop(Gene_field.row)

        # List to store Gene name and pathology
        # We are not using the dictionary for further steps because
        # As we parse other candidate gene files, if the same gene (key)
        # is associated with a different pathology, the existing 
        # gene-pathology pair will be replaced as dictionary cannot 
        # contain redundant keys
        for Gene_identifier in Gene_patho_dict:
            Gene_identifierF = Gene_identifier.split('_')

            # Gene_identifierF[0] -> Gene name

            for ENSG in ENSG_Gene_dict.keys():
                if Gene_identifierF[0] == ENSG_Gene_dict[ENSG]:
                    Gene = ENSG
            
            Pathology = Gene_patho_dict[Gene_identifier]

            # Check if the Gene exists in CandidateGene_dict
            # Happens when same gene is associated with different pathology
            # If Gene exists, then append the new pathology to the list of pathologies
            if CandidateGene_dict.get(Gene, False):
                # Avoid adding same pathology more than once
                if not Pathology in CandidateGene_dict[Gene]:
                    CandidateGene_dict[Gene].append(Pathology)
            else:
                CandidateGene_dict[Gene] = [Pathology]
            
    return CandidateGene_dict

###########################################################

# Parses the dictionary {CandidateGene_dict}
# returned by the function: CandidateGeneParser
#
# Returns a list containing all the pathologies/Phenotypes
def getPathologies(CandidateGene_dict):

    # List of pathologies
    pathologies_list = []

    for candidateGene in CandidateGene_dict:
        for pathology in CandidateGene_dict[candidateGene]:
            if not pathology in pathologies_list:
                pathologies_list.append(pathology)

    return pathologies_list

###########################################################

# Parses the CandidateGene_dict & pathologies_list
#
# Counts the total number of candidate genes
# associated with each pathology
#
# Returns a list with the total count candidate genes (for each pathology)
def CountCandidateGenes(CandidateGene_dict, pathologies_list):

    # List for counting total candidate genes
    # associated with each pathology
    pathology_CandidateCount = [0] * len(pathologies_list)

    # Data lines
    for candidateGene in CandidateGene_dict:
        for pathology in CandidateGene_dict[candidateGene]:
            for i in range(len(pathologies_list)):
                if pathology == pathologies_list[i]:
                    pathology_CandidateCount[i] += 1

    return pathology_CandidateCount

###########################################################

# Parses the UniProt Primary Accession file produced by Uniprot_parser.py
# Required columns are: 'Primary_AC' and 'ENSGs' (can be in any order,
# but they MUST exist)
#
# Also parses the dictionary ENSG_Gene_dict
# returned by the function ENSG_Gene
#
# Maps UniProt Primary Accession to ENSG
# Returns the count of UniProt accessions with unique ENSGs
#
# Count corresponds to total human genes
# This count is later used for calculating
# Benjamini-Hochberg adjusted P-values
def Uniprot_ENSG(inPrimAC, ENSG_Gene_dict):

    UniprotPrimAC_File = open(inPrimAC)

    # Grabbing the header line
    UniprotPrimAC_header = UniprotPrimAC_File.readline()

    UniprotPrimAC_header = UniprotPrimAC_header.rstrip('\n')

    UniprotPrimAC_header_fields = UniprotPrimAC_header.split('\t')

    # Check the column header and grab indexes of our columns of interest
    (UniProt_PrimAC_index, ENSG_index) = (-1, -1)

    for i in range(len(UniprotPrimAC_header_fields)):
        if UniprotPrimAC_header_fields[i] == 'Primary_AC':
            UniProt_PrimAC_index = i
        elif UniprotPrimAC_header_fields[i] == 'ENSGs':
            ENSG_index = i

    if not UniProt_PrimAC_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'Primary_AC' in the file: %s \n" % inPrimAC)
        sys.exit()
    elif not ENSG_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'ENSG' in the file: %s \n" % inPrimAC)
        sys.exit()
    # else grabbed the required column indexes -> PROCEED

    # Compiling regular expression

    # Eliminating Mouse ENSGs
    re_ENSMUST = re.compile('^ENSMUSG')

    # Counter for accessions with single canonical human ENSG
    Count_UniqueENSGs = 0

    # Data lines
    for line in UniprotPrimAC_File:
        line = line.rstrip('\n')
        UniprotPrimAC_fields = line.split('\t')

        # ENSG column  - This is a single string containing comma-seperated ENSGs
        # So we split it into a list that can be accessed later
        UniProt_ENSGs = UniprotPrimAC_fields[ENSG_index].split(',')

        # Initializing empty lists
        human_ENSGs = []
        canonical_human_ENSGs = []

        # Eliminating Mouse ENSGs
        for UniProt_ENSG in UniProt_ENSGs:
            if not re_ENSMUST.match(UniProt_ENSG):
                human_ENSGs.append(UniProt_ENSG)
        if not human_ENSGs:
            continue

        # If ENSG is in the canonical transcripts file
        # Append it to canonical_human_ENSGs
        for ENSG in human_ENSGs:
            if ENSG in ENSG_Gene_dict.keys():
                canonical_human_ENSGs.append(ENSG)

        # Keeping the count of protein with single ENSGs
        if len(canonical_human_ENSGs) == 1:
            Count_UniqueENSGs += 1

    return Count_UniqueENSGs


###########################################################

# Parses the High-quality Interactome produced by Interactome.py
#
# Required columns:
# First column - ENSG of Protein A
# Second column - ENSG of Protein B
#
# Returns 2 dictionaries and 1 list:
# First dictionary contains:
# - key: Protein A; Value: List of interactors
# Second dictionary contains:
# key: Protein B; Value: List of interactors
# These dictionaries are later used to determine
# the no. of interactors for a given protein/gene
#
# The list contains all the interacting proteins from the interactome
def Interacting_Proteins(inInteractome):

    # Input - Interactome file
    Interactome_File = open(inInteractome)

    # Dictionaries to store interacting proteins
    # In ProtA_dict, key -> Protein A; Value -> Protein B
    # In ProtB_dict, key -> Protein B; Value -> Protein A
    ProtA_dict = {}
    ProtB_dict = {}

    # List of all interactors from the Interactome
    All_Interactors_list = []

    # Data lines
    for line in Interactome_File:
        line = line.rstrip('\n')

        Interactome_fields = line.split('\t')

        if Interactome_fields[0] != Interactome_fields[1]:
            # Check if the Key(ProtA) exists in ProtA_dict
            # If yes, then append the interctor to 
            # the list of values (Interactors)
            if ProtA_dict.get(Interactome_fields[0], False):
                ProtA_dict[Interactome_fields[0]].append(Interactome_fields[1])
            else:
                ProtA_dict[Interactome_fields[0]] = [Interactome_fields[1]]

            # Check if the Key(ProtB) exists in ProtB_dict
            # If yes, then append the interctor to 
            # the list of values (Interactors)
            if ProtB_dict.get(Interactome_fields[1], False):
                ProtB_dict[Interactome_fields[1]].append(Interactome_fields[0])
            else:
                ProtB_dict[Interactome_fields[1]] = [Interactome_fields[0]]    

            # Storing all the interactors in All_Interactors_list
            if not Interactome_fields[0] in All_Interactors_list:
                All_Interactors_list.append(Interactome_fields[0])
            elif not Interactome_fields[1] in All_Interactors_list:
                All_Interactors_list.append(Interactome_fields[1])
        # else:
            # NOOP -> The interaction is a self-interaction

    # Closing the file
    Interactome_File.close()

    return ProtA_dict, ProtB_dict, All_Interactors_list

###########################################################

# Parses the dictionaries & list returned
# by the function: Interacting_Proteins
#
# Checks the number of interactors for each gene
# Checks the number of known interactors using the 
# candidateGene_out_list returned by the function: CandidateGene2ENSG
#
# Returns a Dictionary
# Key -> Gene name; 
# Value -> A List of Interactome data associated with each pathology
# For each pathology, following Interactome data is added:
# - Known Interactors count
# - list of Known Interactors
# - P-value
def Interactors_PValue(ProtA_dict, ProtB_dict, All_Interactors_list, CandidateGene_dict, pathologies_list, pathology_CandidateCount, ENSG_Gene_dict, Count_UniqueENSGs):

    # Dictionary to store Gene and
    # Interactome data associated with each pathology
    # Key-> Gene; 
    # Value -> A List of Interactome data associated 
    # with each pathology
    Gene_IntAllpatho = {}

    # Checking the number of interactors for each gene
    for ENSG_index in range(len(All_Interactors_list)):

        # List to store Gene name and interactome
        # data associated with each pathology
        # For each pathology, following Interactome 
        # data is added:
        # - Known Interactors count
        # - list of Known Interactors
        # - P-value
        Gene_AllPatho = []

        Gene_AllPatho.append(All_Interactors_list[ENSG_index])

        # List of interactors
        Interactors = []

        # If Protein is the first protein
        if (All_Interactors_list[ENSG_index] in ProtA_dict.keys()):
            # Get the interacting protein
            for Interactor in ProtA_dict[All_Interactors_list[ENSG_index]]:
                if not Interactor in Interactors:
                    Interactors.append(Interactor)
                    
        # If Protein is the Second protein
        if (All_Interactors_list[ENSG_index] in ProtB_dict.keys()):
            # Get the interacting protein
            for Interactor in ProtB_dict[All_Interactors_list[ENSG_index]]:
                if not Interactor in Interactors:
                    Interactors.append(Interactor)

        for i in range(len(pathologies_list)):

            # List for known interactor(s)
            Known_Interactors = []

            # Initializing a list to store data for each pathology
            Output_eachPatho = []

            # Checking if the interactor is a known ENSG (candidate ENSG)
            for interactor in Interactors:
                if interactor in CandidateGene_dict.keys():
                    for pathology in CandidateGene_dict[interactor]:
                        if pathology == pathologies_list[i]:
                            Known_Interactors.append(interactor)

            # Getting the Gene name for Known Interactors
            for Known_InteractorIndex in range(len(Known_Interactors)):
                Known_Interactors[Known_InteractorIndex] = ENSG_Gene_dict[Known_Interactors[Known_InteractorIndex]]

            if Known_Interactors:
                # Applying Fisher's exact test to calculate p-values
                ComputePvalue_data = [[len(Known_Interactors), len(Interactors)],[pathology_CandidateCount[i], Count_UniqueENSGs]]
                (odds_ratio, p_value) = stats.fisher_exact(ComputePvalue_data)
            # If there are no Known Interactors, 
            # there is no point is computing P-value,
            # So we assign P-value as 1
            else:
                p_value = 1

            if Known_Interactors:
                # Storing Known Interactors as a single comma seperated string
                Known_InteractorsStr = ','.join(Known_Interactor for Known_Interactor in Known_Interactors)
                Output_eachPatho = [len(Known_Interactors), Known_InteractorsStr, p_value]
            else:
                Output_eachPatho = [len(Known_Interactors), '', p_value]

            for data in Output_eachPatho:
                Gene_AllPatho.append(data)

        #  Storing it in the dictionary 
        Gene_IntAllpatho[Gene_AllPatho[0]] = Gene_AllPatho[1:len(Gene_AllPatho)]

    return Gene_IntAllpatho

###########################################################

# Function takes args
# - Canonical transcripts file
# - Candidate Gene(s) file
# - UniProt Primary Accession file
# - Interactome file
#
# Reads on STDIN a TSV file as produced by 5.1_addGTEX.pl
#
# Print to stdout a TSV file with added columns holding
# Interactome data (non-clustering approach)
#
# New columns containing Interactome data are added
# immediately after the 'SYMBOL' column
# This is checked using the Symbol_index
def addInteractome(args):

    # Calling the functions
    ENSG_Gene_dict = ENSG_Gene(args.inCanonicalFile)
    CandidateGene_dict = CandidateGeneParser(args.inCandidateFile, ENSG_Gene_dict)
    pathologies_list = getPathologies(CandidateGene_dict)
    pathology_CandidateCount = CountCandidateGenes(CandidateGene_dict, pathologies_list)
    Count_UniqueENSGs = Uniprot_ENSG(args.inPrimAC, ENSG_Gene_dict)
    (ProtA_dict, ProtB_dict, All_Interactors_list) = Interacting_Proteins(args.inInteractome)
    Gene_IntAllpatho = Interactors_PValue(ProtA_dict, ProtB_dict, All_Interactors_list, CandidateGene_dict, pathologies_list, pathology_CandidateCount, ENSG_Gene_dict, Count_UniqueENSGs)

    # Take STDIN file as produced by 5.1_addGTEX.pl
    addGTEX_output = sys.stdin

    # Get header line
    addGTEX_out_headerLine = addGTEX_output.readline()

    addGTEX_out_headerLine = addGTEX_out_headerLine.rstrip('\n')

    # Stores headers in a list
    addGTEX_out_headers = addGTEX_out_headerLine.split('\t')

    # Check the column header and grab index of our column of interest
    # Here, we want the grab the indices of columns 'SYMBOL' and 'Gene'
    (Symbol_index, Gene_index) = (-1,-1)

    for i in range(len(addGTEX_out_headers)):
        if addGTEX_out_headers[i] == 'SYMBOL':
            Symbol_index = i
        elif addGTEX_out_headers[i] == 'Gene': 
            Gene_index = i    

    if not Symbol_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'SYMBOL' in STDIN")
        sys.exit()
    elif not Gene_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'Gene' in STDIN")    
        sys.exit()
    # else grabbed the required column index -> PROCEED

    # Patho_header_list is a list containing sublists
    # One sublist per pathology (COHORT)
    # Each Sublist contains following header names:
    # - patho_INTERACTORS_COUNT
    # - patho_INTERACTORS
    # - patho_INTERACTORS_PVALUE
    Patho_header_list = [[patho+'_INTERACTORS_COUNT', patho+'_INTERACTORS', patho+'_INTERACTORS_PVALUE'] for patho in pathologies_list]

    # Initializing a list to store all
    # all the elements of a sublist in a
    # single list
    Patho_headers = []

    for headers in Patho_header_list:
        for header in headers:
            Patho_headers.append(header)

    # Adding Interactome data headers to
    # the header of STDIN file as
    # immediately next to the 'SYMBOL' column
    for headerIndex in range(len(Patho_headers)):
        addGTEX_out_headers.insert(headerIndex + (Symbol_index+1), Patho_headers[headerIndex])

    # Printing header
    print('\t'.join(header for header in addGTEX_out_headers))

    # Data lines
    # Add Interactome data and print to STDOUT
    for line in addGTEX_output:
        line = line.rstrip('\n')
        line_fields = line.split('\t')

        # Checking if the Gene present in addGTEX_output exists 
        # as a key in the dictionary: Gene_IntAllpatho
        # If yes, then adding the interactome data associated
        # with this particular gene
        if line_fields[Gene_index] in Gene_IntAllpatho.keys():
            # Inserting Interactome data immediately after the 'SYMBOL' column
            line_fields[Symbol_index+1:Symbol_index+1] = Gene_IntAllpatho[line_fields[Gene_index]]
            print('\t'.join(str(data) for data in line_fields))
        else:
            # If the gene is not present in addGTEX_output, then we assign the same
            # value as when no known interactors were found for a given gene
            # i.e 
            # INTERACTORS_COUNT = 0
            # INTERACTORS = ''
            # INTERACTORS_PVALUE = 1
            line_fields[Symbol_index+1:Symbol_index+1] = [0,'',1] * len(pathologies_list)
            print('\t'.join(str(data) for data in line_fields))  

    # Closing the file
    addGTEX_output.close()

    logging.info("ALL DONE, completed successfully!")

    return

###########################################################

# Taking and handling command-line arguments
def main():
    file_parser = argparse.ArgumentParser(description =
    """

Parse on STDIN a TSV file as produced by 5.1_addGTEX.pl. Also parses the files provided to the arguments. Print to STDOUT a TSV file (similar to the one produced by 5.1_addGTEX.pl) with additional columns holding Interactome data (non-clustering approach)

Arguments [defaults] -> Can be abbreviated to shortest unambiguous prefixes
    """,
    formatter_class = argparse.RawDescriptionHelpFormatter)

    required = file_parser.add_argument_group('Required arguments')
    optional = file_parser.add_argument_group('Optional arguments')

    required.add_argument('--inPrimAC', metavar = "Input File", dest = "inPrimAC", help = 'Uniprot Primary Accession File generated by the UniProt_parser.py', required = True)
    required.add_argument('--inCandidateFile', metavar = "Input File", dest = "inCandidateFile", nargs = '+', help = 'Candidate Genes Input File name(.xlsx)', required = True)
    required.add_argument('--inCanonicalFile', metavar = "Input File", dest = "inCanonicalFile", help = 'Canonical Transcripts file (.gz or non .gz)', required = True)
    required.add_argument('--inInteractome', metavar = "Input File", dest = "inInteractome", help = 'Input File Name (High-quality Human Interactome(.tsv) produced by Build_Interactome.py)', required = True)

    args = file_parser.parse_args()
    addInteractome(args)

if __name__ == "__main__":
    # Logging to Standard Error
    logging.basicConfig(format = "%(levelname)s %(asctime)s: %(filename)s - %(message)s", datefmt='%Y-%m-%d %H:%M:%S', stream = sys.stderr, level = logging.DEBUG)
    logging.addLevelName(logging.INFO, 'I' )
    logging.addLevelName(logging.ERROR, 'E')
    logging.addLevelName(logging.WARNING, 'W')
    main()
